from pathlib import Path
import openpyxl 

import os
from crawler import crawler_org 
from common import common_method as cmn
from const import all_const as cst


# Excelから対象データを抜き出す
def get_control_data(file_path:str) -> list:
    cmn.write_log('get_control_data() 開始')
    wb = openpyxl.load_workbook(file_path,data_only=True)

    row = 2     # 開始行
    lst = []
    for ws in wb.worksheets:

        # シート名の先頭時に「_」があれば除外
        if ws.title[:1] == '_' : continue

        while ws['A'+ str(row)].value != None :
            lst.append({
                'row_number':  row,
                'sheet_name':   ws.title,
                'control':  ws['A'+ str(row)].value,
                'particular': ws['B'+ str(row)].value,
                'object_name': ws['C'+ str(row)].value,
                'handling': ws['D'+ str(row)].value,
                'value' : ws['E'+ str(row)].value 
            })
            row += 1
    cmn.write_log('get_control_data() 終了')
    return lst

   
# -------------------------------------------
# メイン
# -------------------------------------------
def main() :
    try:
        cmn.write_log("main() プログラム開始")

        # Excel チェック
        folder = Path(cst.EXCEL_DIR)
        if not folder.exists() :
            cmn.write_log(f"対象フォルダ({cst.EXCEL_DIR})が存在しない")
            cmn.write_log("プログラム終了")
            exit()

        if not len(list(folder.glob('*.xlsx'))) :
            cmn.write_log(f"対象フォルダ({cst.EXCEL_DIR})にExcelが存在しない")
            cmn.write_log("プログラム終了")
            exit()

        # Excel 開く
        for xls in folder.glob('*.xlsx'):
            if '\~$' in str(xls) : continue  # Excel一時ファイルは対象外
            cmn.write_log( f"[{xls}]を起動しました。")
            lst = get_control_data(xls)

            # ブラウザ動作
            crw = crawler_org()
            crw.run_brawser(lst)


    except Exception as e:
        print(e)
        cmn.write_log("main() Exception検知")
        cmn.write_log("エラーになりました")
    else :
        cmn.write_log("main() プログラム終了")


# 初期動作
if __name__ == '__main__' :
    main()